# Demonstrate python program to read the data from the spreadsheet and write the data in to the spreadsheet

import openpyxl

def read_data_from_spreadsheet(filename):
  workbook = openpyxl.load_workbook(filename)
  sheet = workbook.active

  data = []
  for row in sheet.iter_rows(values_only=True):
    data.append(row)

  return data

def write_data_to_spreadsheet(filename, data):
  workbook = openpyxl.Workbook()
  sheet = workbook.active

  for row in data:
    sheet.append(row)

  workbook.save(filename)


input_filename = 'input.xlsx'

read_data = read_data_from_spreadsheet(input_filename)

print("data read from spreadsheet:")
for row in read_data:
  print(row)

write_data = [
    ['name','age','city'],
    ['anil',45,'tumkur'],
    ['sunil',37,'bangalore'],
    ['hari',43,'gubbi']
]
output_filename = 'output.xlsx'
write_data_to_spreadsheet(output_filename,write_data)
print("Data written to spreadsheet:",output_filename)
